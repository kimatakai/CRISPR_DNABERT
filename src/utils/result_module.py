

import os
import tqdm
import yaml
from itertools import combinations
import numpy as np
import pandas as pd
from openpyxl import Workbook, load_workbook
from sklearn.metrics import roc_auc_score, precision_recall_curve, auc, matthews_corrcoef, accuracy_score, recall_score, precision_score, f1_score
from scipy.stats import wilcoxon

import models.data_loader as data_loader
import models.result as result
import visualization.plot_result_fig as plot_result_fig


class ResultsBaseClass:
    def __init__(self, config: dict) -> None:
        self.config = config
        self.models_name_list = config["model_info"]["models_name"]
        self.dataset_name = config["dataset_name"]["dataset_current"]
        self.folds_list = config["folds"].copy()
        self.iters_list = config["iters"].copy()
        self.exe_type = config["exe_type"]
        self.with_epigenetic = config["with_epigenetic"]
        
        # Paths information
        self.probability_paths = config["probability_paths"]
        self.result_paths = config["result_paths"]
        
        self.metrics = ["accuracy", "recall", "precision", "f1_score", "mcc", "roc_auc", "pr_auc", "confusion_matrix"]
        self.confusion_matrix_labels = ["TN", "FP", "FN", "TP"]
        self.confusion_matrix_indices = {"TN": (0, 0), "FP": (0, 1), "FN": (1, 0), "TP": (1, 1)}
        self.confusion_matrix_indices_rev = {(0, 0): "TN", (0, 1): "FP", (1, 0): "FN", (1, 1): "TP"}
        self.mm_blg_type_labels = ["all"] + [(mm, blg) for mm in range(7) for blg in range(2)]
    
    def load_dataset_info(self) -> dict:
        aggregated_label = []
        aggregated_mismatch = []
        aggregated_bulge = []
        self.config["fold"] = "all"
        self.config["iter"] = 0
        DataLoaderClass = data_loader.DataLoaderClass(self.config)
        DataLoaderClass.load_sgrna_list()
        dataset_dict = DataLoaderClass.load_and_convert_to_dict()
        for fold in tqdm.tqdm(self.folds_list, total=len(self.folds_list), desc="Loading dataset information"):
            for iter in self.iters_list:
                self.config["fold"] = fold
                _dataset_dict = DataLoaderClass.split_dataset(dataset_dict, fold)
                aggregated_label += _dataset_dict["label"]
                aggregated_mismatch += _dataset_dict["mismatch"]
                aggregated_bulge += _dataset_dict["bulge"]
        return {
            "label": np.array(aggregated_label),
            "mismatch": np.array(aggregated_mismatch),
            "bulge": np.array(aggregated_bulge)
        }
    
    def load_result(self, result_path: str) -> dict:
        if not os.path.exists(result_path):
            raise FileNotFoundError(f"Result file not found: {result_path}")
        with open(result_path, 'r') as f:
            result = yaml.safe_load(f)
        for metric in self.metrics:
            if metric not in result:
                raise ValueError(f"Incorrect result format: {result_path}")
        return result
    
    def load_probability(self, prob_path: str) -> np.ndarray:
        if not os.path.exists(prob_path):
            raise FileNotFoundError(f"Probability file not found: {prob_path}")
        prob_array = np.load(prob_path)
        return prob_array
    
    def aggregate_results(self) -> dict:
        aggregated_results = {}
        for model_name in tqdm.tqdm(self.models_name_list, total=len(self.models_name_list), desc="Aggregating results"):
            model_results = {metric: [] for metric in self.metrics}
            for path in self.result_paths[model_name]:
                result = self.load_result(path)
                for metric in self.metrics:
                    model_results[metric].append(result[metric])
            aggregated_results[model_name] = model_results
        return aggregated_results
    
    def aggregate_probabilities(self) -> dict:
        aggregated_probabilities = {}
        for model_name in tqdm.tqdm(self.models_name_list, total=len(self.models_name_list), desc="Aggregating probabilities"):
            model_probs = []
            for path in self.probability_paths[model_name]:
                prob_array = self.load_probability(path)
                model_probs.append(prob_array)
            aggregated_probabilities[model_name] = np.concatenate(model_probs)
        return aggregated_probabilities

    def aggregate_confusion_matrices(self, aggregated_probabilities: dict, aggregated_labels: dict, aggregated_mismatches: dict, aggregated_bulges: dict) -> dict:
        aggregated_confusion = {l: {model_name: {cm_l: 0 for cm_l in self.confusion_matrix_labels} for model_name in self.models_name_list} for l in self.mm_blg_type_labels}
        for model_name in self.models_name_list:
            aggregated_predictions = (aggregated_probabilities[model_name] >= 0.5).astype(int)
            for prediction, true_label, mismatch, bulge in zip(aggregated_predictions, aggregated_labels, aggregated_mismatches, aggregated_bulges):
                cm_label = self.confusion_matrix_indices_rev[(true_label, prediction)]
                aggregated_confusion["all"][model_name][cm_label] += 1
                aggregated_confusion[(mismatch, bulge)][model_name][cm_label] += 1
        return aggregated_confusion
    
    def wilcoxon_signed_rank_test(self, aggregated_results: dict, model_a: str, model_b: str) -> dict: # {"metric": p_value}
        if model_a not in self.models_name_list or model_b not in self.models_name_list:
            raise ValueError(f"Models {model_a} or {model_b} not in the model list.")
        results_a = aggregated_results[model_a]
        results_b = aggregated_results[model_b]
        p_values = {}
        for metric in [m for m in self.metrics if m != "confusion_matrix"]:
            if metric in results_a and metric in results_b:
                stat, p_value = wilcoxon(results_a[metric], results_b[metric], alternative="two-sided")
                p_values[metric] = p_value
        return p_values

    def adjust_p_values_bh(self, p_values: list) -> list:
        n = len(p_values)
        sorted_indices = np.argsort(p_values)
        sorted_p_values = np.array(p_values)[sorted_indices]
        adjusted_p_values = np.empty(n)
        cumulative_min = 1.0
        for i in range(n - 1, -1, -1):
            rank = i + 1
            adjusted_p = sorted_p_values[i] * n / rank
            cumulative_min = min(cumulative_min, adjusted_p)
            adjusted_p_values[i] = cumulative_min
        adjusted_p_values_corrected = np.empty(n)
        adjusted_p_values_corrected[sorted_indices] = adjusted_p_values
        return adjusted_p_values_corrected.tolist()
    
    def aggregate_wilcoxon_p_values(self, aggregated_results: dict) -> dict:
        p_values = {metric: [] for metric in self.metrics if metric != "confusion_matrix"}
        adjusted_p_values_dict = {}
        for i_a, model_a in tqdm.tqdm(enumerate(self.models_name_list), total=len(self.models_name_list), desc="Calculating Wilcoxon p-values"):
            for model_b in [m for m in self.models_name_list[i_a + 1:] if m != model_a]:
                p_values_dict = self.wilcoxon_signed_rank_test(aggregated_results, model_a, model_b)
                for metric, p_value in p_values_dict.items():
                    p_values[metric].append(p_value)
                adjusted_p_values_dict[(model_a, model_b)] = {}
        # Adjust p-values using Benjamini-Hochberg
        for metric, p_vals in p_values.items():
            i = 0
            adj_p_values = self.adjust_p_values_bh(p_vals) # List
            for i_a, model_a in enumerate(self.models_name_list):
                for model_b in [m for m in self.models_name_list[i_a + 1:] if m != model_a]:
                    adjusted_p_values_dict[(model_a, model_b)][metric] = adj_p_values[i]
                    i += 1
        return adjusted_p_values_dict

    def aggregation_time(self) -> dict:
        if self.dataset_name not in ["Lazzarotto_2020_CHANGE_seq", "Lazzarotto_2020_GUIDE_seq", "SchmidBurgk_2020_TTISS"]:
            raise ValueError("No time paths available in the configuration.")
        aggregated_times = {model_name: [] for model_name in self.models_name_list}
        for model_name in self.models_name_list:
            for path in self.config["time_paths"][model_name]:
                if not os.path.exists(path):
                    raise FileNotFoundError(f"Time file not found: {path}")
                with open(path, 'r') as f:
                    exe_time = float(f.read().strip())
                    aggregated_times[model_name].append(exe_time)
        return aggregated_times


class ResultsWorkbook(ResultsBaseClass):
    def __init__(self, config: dict) -> None:
        super().__init__(config)
        # Path information
        self.excel_file_path = config["excel_path"]
        # Aggregated results
        self.aggregated_results = self.aggregate_results() # Dict
        self.aggregated_p_values = self.aggregate_wilcoxon_p_values(self.aggregated_results) # Dict
        self.aggregated_probabilities = self.aggregate_probabilities() # Dict
        self.aggregated_dataset_info = self.load_dataset_info() # Dict
        self.aggregated_labels = self.aggregated_dataset_info["label"]
        self.aggregated_mismatches = self.aggregated_dataset_info["mismatch"]
        self.aggregated_bulges = self.aggregated_dataset_info["bulge"]
        self.aggregated_cm = self.aggregate_confusion_matrices(self.aggregated_probabilities, self.aggregated_labels, self.aggregated_mismatches, self.aggregated_bulges)

    def create_workbook(self):
        os.makedirs(os.path.dirname(self.excel_file_path), exist_ok=True)
        wb = Workbook()
        ws = wb.active
        ws.title = "File information"
        data = [
            ["Dataset", self.dataset_name],
            ["Model"] + self.models_name_list,
            ["Fold"] + self.folds_list,
            ["Experiment iteration"] + self.iters_list,
            ["Training method", self.exe_type],
            ["With epigenetic feature?", self.with_epigenetic]
        ]
        for row in data:
            ws.append(row)
        wb.save(self.excel_file_path)
    
    def add_result_aggregation_sheet(self):
        # Add result aggregation sheet (mean, median, std, max, min)
        wb = load_workbook(self.excel_file_path)
        ws_result_agg = wb.create_sheet(title="Result Aggregation")
        _metrics = [metrics for metrics in self.metrics if metrics != "confusion_matrix"]

        # Create Sheet Data
        sheet_data = []
        sheet_data.append(["Results Aggregation"])
        sheet_data.append([])
        for model_name in self.models_name_list:
            sheet_data.append([model_name])
            metrics_list, mean_list, median_list, std_list, max_list, min_list = ["Metrics"], ["Mean"], ["Median"], ["Std"], ["Max"], ["Min"]
            for metrics in _metrics:
                metrics_list.append(metrics)
                res_array = np.array(self.aggregated_results[model_name][metrics])
                mean_list.append(np.mean(res_array))
                median_list.append(np.median(res_array))
                std_list.append(np.std(res_array))
                max_list.append(np.max(res_array))
                min_list.append(np.min(res_array))
            sheet_data += [metrics_list, mean_list, median_list, std_list, max_list, min_list]
            sheet_data.append([])
        # Write to Sheet
        for row in sheet_data:
            ws_result_agg.append(row)
        wb.save(self.excel_file_path)
    
    def add_wilcoxon_pvalue_sheet(self):
        wb = load_workbook(self.excel_file_path)
        ws_wilcoxon = wb.create_sheet(title="Wilcoxon P-Values")
        _metrics = [metrics for metrics in self.metrics if metrics != "confusion_matrix"]
        
        # Create Sheet Data
        sheet_data = []
        sheet_data.append(["Wilcoxon Signed-Rank Test P-Values"])
        sheet_data.append([])
        for metric in _metrics:
            sheet_data.append(["metric", metric])
            sheet_data.append([""] + self.models_name_list)
            for model_a in self.models_name_list:
                row = [model_a]
                for model_b in self.models_name_list:
                    if model_a == model_b:
                        row.append("-")
                    else:
                        p_value = self.aggregated_p_values.get((model_a, model_b), {}).get(metric, "N/A")
                        row.append(p_value)
                sheet_data.append(row)
            sheet_data.append([])
        # Write to Sheet
        for row in sheet_data:
            ws_wilcoxon.append(row)
        wb.save(self.excel_file_path)
    
    def add_confusion_matrix_sheet(self):
        wb = load_workbook(self.excel_file_path)
        ws_cm = wb.create_sheet(title="Confusion Matrices")
        # Create Sheet Data
        sheet_data = []
        sheet_data.append(["Confusion Matrices"])
        sheet_data.append([])
        for mm_blg in self.mm_blg_type_labels:
            for model_name in self.models_name_list:
                if mm_blg == "all":
                    sheet_data.append([f"All data", f"Model: {model_name}"])
                else:
                    sheet_data.append([f"Mismatch: {mm_blg[0]}", f"Bulge: {mm_blg[1]}", f"Model: {model_name}"])
                sheet_data.append([""] + self.confusion_matrix_labels)
                sheet_data.append(["Count"] + [self.aggregated_cm[mm_blg][model_name][cm_l] for cm_l in self.confusion_matrix_labels])
                sheet_data.append([])
        # Write to Sheet
        for row in sheet_data:
            ws_cm.append(row)
        wb.save(self.excel_file_path)
    
    def add_execution_time_sheet(self):
        if self.dataset_name not in ["Lazzarotto_2020_CHANGE_seq", "Lazzarotto_2020_GUIDE_seq", "SchmidBurgk_2020_TTISS"]:
            print("No execution time data available.")
            return
        wb = load_workbook(self.excel_file_path)
        ws_time = wb.create_sheet(title="Execution Time")
        aggregated_times = self.aggregation_time()
        
        # Create Sheet Data
        sheet_data = []
        sheet_data.append(["Execution Time (seconds)"])
        sheet_data.append([])
        for model_name in self.models_name_list:
            sheet_data.append([model_name])
            times = aggregated_times[model_name]
            sheet_data.append(["", "Mean", "Median", "Std", "Max", "Min"])
            sheet_data.append(["Time (s)", np.mean(times), np.median(times), np.std(times), np.max(times), np.min(times)])
            sheet_data.append([])
        # Write to Sheet
        for row in sheet_data:
            ws_time.append(row)
        wb.save(self.excel_file_path)


class ResultsForEnsemble(ResultsBaseClass):
    def __init__(self, config: dict) -> None:
        self.config = config
        self.models_name_list = config["model_info"]["models_name"]
        self.dataset_name = config["dataset_name"]["dataset_current"]
        self.folds_list = config["folds"].copy()
        self.iters_list = config["iters"].copy()
        self.exe_type = config["exe_type"]
        self.with_epigenetic = config["with_epigenetic"]
        self.probability_paths = config["probability_paths"]
        self.result_paths = config["result_paths"]
        # Path information
        self.excel_file_path = config["excel_path"]
        self.fig_dir_path = config["fig_dir_path"]
        # Metrics information
        self.metrics = ["accuracy", "recall", "precision", "f1_score", "mcc", "roc_auc", "pr_auc", "confusion_matrix"]
        self.confusion_matrix_labels = ["TN", "FP", "FN", "TP"]
        self.confusion_matrix_indices = {"TN": (0, 0), "FP": (0, 1), "FN": (1, 0), "TP": (1, 1)}
        self.confusion_matrix_indices_rev = {(0, 0): "TN", (0, 1): "FP", (1, 0): "FN", (1, 1): "TP"}
        # Aggregated probabilities
        self.aggregated_probabilities = self.aggregate_probabilities_for_ensemble() # Dict
        self.ensemble_probabilities = self.aggregate_ensemble_probabilities(self.aggregated_probabilities) # Dict {ensemble_name: {(fold, iter): prob_array}}
        self.aggregated_labels = self.load_label_info_for_ensemble() # Dict {fold: label_array}
        self.ensemble_results = self.aggregate_ensemble_results(self.ensemble_probabilities, self.aggregated_labels) # Dict {ensemble_name: {(fold, iter): metrics_dict}}
        self.ensemble_results_list = self.aggregate_ensemble_results_to_list() # Dict {ensemble_name: {metric: [values]}}
        self.find_max_ensemble_result() # self.max_results
        

    def aggregate_probabilities_for_ensemble(self) -> dict:
        aggregated_probabilities = {}
        for model_name in tqdm.tqdm(self.models_name_list, total=len(self.models_name_list), desc="Aggregating probabilities for ensemble"):
            aggregated_probabilities[model_name] = {}
            i = 0
            for fold in self.folds_list:
                for iter in self.iters_list:
                    path = self.probability_paths[model_name][i]
                    prob_array = self.load_probability(path)
                    aggregated_probabilities[model_name][(fold, iter)] = prob_array
                    i += 1
        return aggregated_probabilities

    def load_label_info_for_ensemble(self) -> dict:
        aggregated_label = {}
        self.config["fold"] = "all"
        self.config["iter"] = 0
        DataLoaderClass = data_loader.DataLoaderClass(self.config)
        DataLoaderClass.load_sgrna_list()
        dataset_dict = DataLoaderClass.load_and_convert_to_dict()
        for fold in tqdm.tqdm(self.folds_list, total=len(self.folds_list), desc="Loading dataset information"):
            self.config["fold"] = fold
            _dataset_dict = DataLoaderClass.split_dataset(dataset_dict, fold)
            test_idx = _dataset_dict["test_idx"]
            aggregated_label[fold] = np.array(_dataset_dict["label"])[test_idx]
        return aggregated_label

    def aggregate_ensemble_probabilities(self, aggregated_probabilities: dict) -> dict:
        ensemble_probabilities = {}
        combinatioin_models_name_list = []
        for r in range(2, len(self.models_name_list) + 1):
            combinatioin_models_name_list += list(combinations(self.models_name_list, r))
        for combination_tuple in tqdm.tqdm(combinatioin_models_name_list, total=len(combinatioin_models_name_list), desc="Aggregating ensemble probabilities"):
            if len(combination_tuple) == len(self.models_name_list):
                ensemble_name = "Ensemble_All"
            else:
                ensemble_name = combination_tuple
            _models_list_temp = list(combination_tuple)
            ensemble_probabilities[ensemble_name] = {}
            i = 0
            for fold in self.folds_list:
                for iter in self.iters_list:
                    ensemble_prob_ = []
                    for model_name in _models_list_temp:
                        prob = aggregated_probabilities[model_name][(fold, iter)]
                        ensemble_prob_.append(prob)
                    ensemble_prob_ = np.mean(ensemble_prob_, axis=0)
                    ensemble_probabilities[ensemble_name][(fold, iter)] = ensemble_prob_
                    i += 1
        return ensemble_probabilities
    
    def aggregate_ensemble_results(self, ensemble_probabilities: dict, aggregated_labels: dict) -> dict:
        ensemble_results = {}
        combinatioin_models_name_list = []
        for r in range(2, len(self.models_name_list)):
            combinatioin_models_name_list += list(combinations(self.models_name_list, r))
        combinatioin_models_name_list += ["Ensemble_All"]
        for combination_tuple in tqdm.tqdm(combinatioin_models_name_list, total=len(combinatioin_models_name_list), desc="Aggregating ensemble results"):
            ensemble_name = combination_tuple
            ensemble_results[ensemble_name] = {}
            for fold in self.folds_list:
                for iter in self.iters_list:
                    label_array = aggregated_labels[fold]
                    prob_array = ensemble_probabilities[ensemble_name][(fold, iter)]
                    pred_label_array = np.where(prob_array > 0.5, 1, 0)
                    metrics_dict = result.return_metrics(fold, iter, label_array, pred_label_array, prob_array, if_show=False)
                    ensemble_results[ensemble_name][(fold, iter)] = metrics_dict
        return ensemble_results
    
    def find_max_ensemble_result(self) -> dict:
        combinatioin_models_name_list = []
        for r in range(2, len(self.models_name_list)):
            combinatioin_models_name_list += list(combinations(self.models_name_list, r))
        combinatioin_models_name_list += ["Ensemble_All"]
        self.max_results = {"model_name": None, "score": 0}
        for combination_tuple in tqdm.tqdm(combinatioin_models_name_list, total=len(combinatioin_models_name_list), desc="Finding max ensemble result"):
            pr_auc_list = self.ensemble_results_list[combination_tuple]["pr_auc"]
            mean_pr_auc = np.median(pr_auc_list)
            if mean_pr_auc > self.max_results["score"]:
                self.max_results["score"] = mean_pr_auc
                self.max_results["model_name"] = combination_tuple
        print(self.max_results)
    
    def aggregate_ensemble_results_to_list(self) -> dict:
        ensemble_results_list = {}
        combinatioin_models_name_list = []
        for r in range(2, len(self.models_name_list)):
            combinatioin_models_name_list += list(combinations(self.models_name_list, r))
        combinatioin_models_name_list += ["Ensemble_All"]
        for combination_tuple in tqdm.tqdm(combinatioin_models_name_list, total=len(combinatioin_models_name_list), desc="Aggregating ensemble results to list"):
            ensemble_results_list[combination_tuple] = {metric: [] for metric in self.metrics if metric != "confusion_matrix"}
            for fold in self.folds_list:
                for iter in self.iters_list:
                    metrics_dict = self.ensemble_results[combination_tuple][(fold, iter)]
                    for metric in self.metrics:
                        if metric != "confusion_matrix":
                            ensemble_results_list[combination_tuple][metric].append(metrics_dict[metric])
        print(ensemble_results_list["Ensemble_All"]["pr_auc"])
        return ensemble_results_list
        
    def save_ensemble_results(self) -> None:
        best_ensemble_name = self.max_results["model_name"]
        for fold in self.folds_list:
            for iter in self.iters_list:
                prob_array = self.ensemble_probabilities[best_ensemble_name][(fold, iter)]
                result_dict = self.ensemble_results[best_ensemble_name][(fold, iter)]
                # Save probability
                prob_path = self.config["probability_paths_ensemble"][(fold, iter)]
                np.save(prob_path, prob_array)
                # Save result
                result_path = self.config["result_paths_ensemble"][(fold, iter)]
                result.save_results(result_dict, result_path)

    def create_workbook(self):
        os.makedirs(os.path.dirname(self.excel_file_path), exist_ok=True)
        wb = Workbook()
        ws = wb.active
        ws.title = "File information for Ensemble"
        data = [
            ["Dataset", self.dataset_name],
            ["Model"] + self.models_name_list,
            ["Fold"] + self.folds_list,
            ["Experiment iteration"] + self.iters_list,
            ["Training method", self.exe_type],
            ["With epigenetic feature?", self.with_epigenetic]
        ]
        for row in data:
            ws.append(row)
        wb.save(self.excel_file_path)
    
    def add_ensemble_result_sheet(self):
        _combinatioin_models_name_list_temp = []
        for r in range(2, len(self.models_name_list)):
            _combinatioin_models_name_list_temp += list(combinations(self.models_name_list, r))
        combinatioin_models_name_list = [self.max_results["model_name"], "Ensemble_All"]
        combinatioin_models_name_list += [combination_tuple for combination_tuple in _combinatioin_models_name_list_temp if len(combination_tuple) == len(self.models_name_list) - 1]
        for combination_tuple in tqdm.tqdm(combinatioin_models_name_list, total=len(combinatioin_models_name_list), desc="Adding ensemble result sheet"):
            if combination_tuple == self.max_results["model_name"]:
                sheet_name = "Best_Ensemble"
            elif combination_tuple == "Ensemble_All":
                sheet_name = "Ensemble_All"
            else:
                removed_model = [m for m in self.models_name_list if m not in combination_tuple][0]
                sheet_name = f"Ensemble_no_{removed_model}"
            wb = load_workbook(self.excel_file_path)
            ws_ensemble = wb.create_sheet(title=sheet_name)
            # Create Sheet Data
            sheet_data = []
            models = [m for m in combination_tuple] if combination_tuple != "Ensemble_All" else self.models_name_list
            sheet_data.append([f"Ensemble Model:"] + models)
            sheet_data.append([])
            metrics_list, mean_list, median_list, std_list, max_list, min_list = ["Metrics"], ["Mean"], ["Median"], ["Std"], ["Max"], ["Min"]
            for metrics in [m for m in self.metrics if m != "confusion_matrix"]:
                metrics_list.append(metrics)
                res_array = np.array(self.ensemble_results_list[combination_tuple][metrics])
                mean_list.append(np.mean(res_array))
                median_list.append(np.median(res_array))
                std_list.append(np.std(res_array))
                max_list.append(np.max(res_array))
                min_list.append(np.min(res_array))
            sheet_data += [metrics_list, mean_list, median_list, std_list, max_list, min_list]
            sheet_data.append([])
            # Write to Sheet
            for row in sheet_data:
                ws_ensemble.append(row)
            wb.save(self.excel_file_path)

    def plot_ensemble_importance(self):
        pr_auc_score = {}
        _combinatioin_models_name_list_temp = []
        for r in range(2, len(self.models_name_list)):
            _combinatioin_models_name_list_temp += list(combinations(self.models_name_list, r))
        combinatioin_models_name_list = ["Ensemble_All"]
        combinatioin_models_name_list += [combination_tuple for combination_tuple in _combinatioin_models_name_list_temp if len(combination_tuple) == len(self.models_name_list) - 1]
        for combination_tuple in tqdm.tqdm(combinatioin_models_name_list, total=len(combinatioin_models_name_list), desc="Plotting ensemble importance"):
            if combination_tuple == "Ensemble_All":
                score_key = "all"
            else:
                removed_model = [m for m in self.models_name_list if m not in combination_tuple][0]
                score_key = removed_model
            pr_auc_list = self.ensemble_results_list[combination_tuple]["pr_auc"]
            pr_auc_score[score_key] = np.mean(pr_auc_list)
        plot_result_fig.plot_ensemble_importance_barplot(self.config, pr_auc_score, save_dir=self.fig_dir_path, dpi=300)


class ResultForVisualization(ResultsBaseClass):
    def __init__(self, config: dict) -> None:
        super().__init__(config)
        # Aggregated results
        self.aggregated_results = self.aggregate_results() # Dict
        self.aggregated_p_values = self.aggregate_wilcoxon_p_values(self.aggregated_results) # Dict
        self.fig_dir_path = config["fig_dir_path"]
    
    def plot_box_plot(self):
        plot_result_fig.plot_boxplot(self.config, self.aggregated_results, self.aggregated_p_values, self.models_name_list, self.metrics, save_dir=self.fig_dir_path, dpi=300)
        pass
        
        
        
        