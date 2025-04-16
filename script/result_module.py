
import sys
sys.path.append("script/")
import os

import config
import data_loader
import visualize_module
import utilities_module

import tqdm
from itertools import combinations
import numpy as np
import pandas as pd
from openpyxl import Workbook, load_workbook
from sklearn.metrics import roc_auc_score, precision_recall_curve, auc, matthews_corrcoef, accuracy_score, recall_score, precision_score, f1_score
from scipy.stats import wilcoxon



def create_workbooks(model_list: list, fold_list: list, exp_id_list: list, datatype: str, file_name: str):
    if not os.path.exists(config.result_base_dir_path):
        os.makedirs(config.result_base_dir_path)
    wb = Workbook()
    ws = wb.active
    ws.title = "File information"
    data = [
        ["Data type", datatype],
        ["Model"] + model_list,
        ["Fold"] + fold_list,
        ["Experiment ID"] + exp_id_list
    ]
    for row in data:
        ws.append(row)
    wb.save(f"{config.result_base_dir_path}/{file_name}.xlsx")
    return


def all_combinations(input_list):
    """
    Generate all possible combinations of two or more elements from the given list.
    
    Args:
    input_list (list): The list of elements to generate combinations from.
    
    Returns:
    list: A list of tuples, each containing a combination of the elements.
    """
    result = []
    for r in range(2, len(input_list) + 1):
        result.extend(combinations(input_list, r))
    return result


def caluculate_metrics(true_label: np.array, 
                       probabilities: np.array,
                       datatype,
                       metrics_list: list=['accuracy', 'recall', 'precision', 'specificity', 'FPR', 'FDR', 'f1', 'ROC-AUC', 'PR-AUC', 'mcc'], 
                       show: bool=False):
    results = {}
    
    probabilities_label1 = probabilities[:, 1]
    predictions = np.argmax(probabilities, axis=1)
    
    # Metrics
    if "accuracy" in metrics_list:
        accuracy = accuracy_score(true_label, predictions)
        results["accuracy"] = accuracy
    if "recall" in metrics_list:
        recall = recall_score(true_label, predictions)
        results["recall"] = recall
    if "precision" in metrics_list:
        precision = precision_score(true_label, predictions)
        results["precision"] = precision
    if "specificity" in metrics_list:
        specificity = recall_score(true_label, predictions, pos_label=0)
        results["specificity"] = specificity
    if "FPR" in metrics_list:
        fpr = 1 - specificity
        results["FPR"] = fpr
    if "FDR" in metrics_list:
        fdr = 1 - precision
        results["FDR"] = fdr
    if "f1" in metrics_list:
        f1 = f1_score(true_label, predictions)
        results["f1"] = f1
    if "ROC-AUC" in metrics_list:
        auc_score = roc_auc_score(true_label, probabilities_label1)
        results["ROC-AUC"] = auc_score
    if "PR-AUC" in metrics_list:
        p, r, _ = precision_recall_curve(true_label, probabilities_label1)
        prauc_score = auc(r, p)
        results["PR-AUC"] = prauc_score
    if "mcc" in metrics_list:
        mcc = matthews_corrcoef(true_label, predictions)
        results["mcc"] = mcc
    
    if show:
        for metric in results.keys():
            print(f"{metric} : {results[metric]:.4f}")
    
    return results



# define data loader class for result analysis
class DataLoaderClassForResult:
    def __init__(self, model_list: list, fold_list: list, exp_id_list: list, datatype: str, metrics_list: list):
        self.model_list = model_list
        self.fold_list = fold_list
        self.exp_id_list = exp_id_list
        self.datatype = datatype
        self.metrics_list = metrics_list
        self.file_name = f"result_{self.datatype}.xlsx"
        if self.datatype == "changeseq":
            self.dataset_file_path = f"{config.yaish_et_al_data_path}/CHANGEseq/include_on_targets/CHANGEseq_CR_Lazzarotto_2020_dataset.csv"
            self.sgrna_list_path = f"{config.yaish_et_al_data_path}/CHANGEseq_sgRNAs_list.csv"
            self.sgrna_fold_path = f"{config.yaish_et_al_data_path}/CHANGEseq_sgRNAs_folds_split.csv"
        elif self.datatype == "guideseq" or self.datatype == "transfer":
            self.dataset_file_path = f"{config.yaish_et_al_data_path}/GUIDEseq/include_on_targets/GUIDEseq_CR_Lazzarotto_2020_dataset.csv"
            self.sgrna_list_path = f"{config.yaish_et_al_data_path}/GUIDEseq_sgRNAs_list.csv"
            self.sgrna_fold_path = f"{config.yaish_et_al_data_path}/GUIDEseq_sgRNAs_folds_split.csv"
        
        self.model_name_mapping = {
            "dnabert": "DNABERT",
            "dnabert-no-pretrain": "DNABERT-No-Pretrain",
            "dnabert-epi": "DNABERT-Epi",
            "dnabert-epi-ablation": "DNABERT-Epi-Ablation",
            "gru-embed": "GRU-Emb",
            "crispr-bert": "CRISPR-BERT",
            "crispr-hw": "CRISPR-HW",
            "crispr-dipoff": "CRISPR-DIPOFF",
            "crispr-bert-2025": "CrisprBERT",
            "ensemble": "Ensemble"}
        self.metrics_mapping = {
            "accuracy": "Accuracy",
            "recall": "Recall",
            "precision": "Precision",
            "specificity": "Specificity",
            "FPR": "FPR",
            "FDR": "FDR",
            "f1": "F1",
            "ROC-AUC": "ROC-AUC",
            "PR-AUC": "PR-AUC",
            "mcc": "MCC"}
    
    
    def create_workbooks(self):
        if not os.path.exists(config.result_base_dir_path):
            os.makedirs(config.result_base_dir_path)
        wb = Workbook()
        ws = wb.active
        ws.title = "File information"
        data = [
            ["Data type", self.datatype],
            ["Model"] + [self.model_name_mapping[model_name] for model_name in self.model_list],
            ["Fold"] + self.fold_list,
            ["Experiment ID"] + self.exp_id_list
        ]
        for row in data:
            ws.append(row)
        wb.save(f"{config.result_base_dir_path}/{self.file_name}")
        return
    
    def load_dataset(self) -> pd.DataFrame:
        '''
        output:
        pd.DataFrame
        '''
        self.dataset_df = pd.read_csv(self.dataset_file_path)
        self.flag_load_dataset = True
        return self.dataset_df

    def load_test_info(self) -> dict:
        '''
        output:
        dict_type data
        test_sgrna_names, test_sgrna_seqs
        '''
        # sgRNA,chrom,SiteWindow,Align.strand,Align.chromStart,Align.chromEnd,Align.off-target,Align.sgRNA,Align.#Mismatches,Align.#Bulges,reads
        # Load sgRNA information
        sgRNAs_json = data_loader.load_sgrna_name()
        sgrna_seq2name_dict = {sgrna_seq:sgrna_name for sgrna_seq, sgrna_name in zip(sgRNAs_json["sgRNAs_seq"], sgRNAs_json["sgRNAs_name"])}
        fold_sgrna_df = pd.read_csv(self.sgrna_fold_path)

        # Test
        self.fold_sgrna_info_dict = {"test_names_list": {}, "test_seq_list": {}}
        for f in self.fold_list:
            test_sgrna_seq_list = [sgrna_seq for sgrna_seq in fold_sgrna_df.iloc[f].values.tolist() if type(sgrna_seq) == str]
            test_sgrna_name_list = [sgrna_seq2name_dict[sgrna_seq] for sgrna_seq in test_sgrna_seq_list]
            self.fold_sgrna_info_dict["test_names_list"][f] = test_sgrna_name_list
            self.fold_sgrna_info_dict["test_seq_list"][f] = test_sgrna_seq_list
        self.flag_fold_sgrna_info_dict = True
        return self.fold_sgrna_info_dict

    def return_test_label(self) -> dict:
        '''
        output:
        dict_type data
        self.test_label_dict
        if datatype is guideseq -> {fold: label}
        '''
        
        if self.datatype != "tf-trueot":
            if not self.flag_load_dataset:
                self.load_dataset()
            if not self.flag_fold_sgrna_info_dict:
                self.load_test_info()
            self.test_label_dict = {}

            for f in self.fold_list:
                label_list = []
                sgrna_seq_list = self.fold_sgrna_info_dict["test_seq_list"][f]
                for sgrna_seq in sgrna_seq_list:
                    dataset_df_sgrna = self.dataset_df[self.dataset_df["sgRNA"] == sgrna_seq]
                    # Split offtarget or non-offtarget
                    data_df_sgrna_offtarget = dataset_df_sgrna[dataset_df_sgrna["reads"] >= 1]
                    data_df_sgrna_non_offtarget = dataset_df_sgrna[dataset_df_sgrna["reads"] == 0]
                    # Processing for off-target row
                    offtarget_label_array = np.array([1]*data_df_sgrna_offtarget.shape[0], dtype=np.int8)
                    label_list.append(offtarget_label_array)
                    # Processing for non-offtarget row
                    non_offtarget_label_array = np.array([0]*data_df_sgrna_non_offtarget.shape[0], dtype=np.int8)
                    label_list.append(non_offtarget_label_array)
                # Concatenate label array
                label_array = np.concatenate(label_list, axis=0)
                # Add to dict
                self.test_label_dict[f] = label_array
        else:
            self.test_label_dict = {}
            label_array = np.zeros((self.dataset_df.shape[0],), dtype=np.int8)
            for idx, row in enumerate(self.dataset_df.iterrows()):
                label_array[idx] = int(row.label)
            for f in self.fold_list:
                self.test_label_dict[f] = label_array
        
        return self.test_label_dict
    
    def return_mismatch_index_dict(self) -> dict:
        if not self.flag_load_dataset:
            self.load_dataset()
        if not self.flag_fold_sgrna_info_dict:
            self.load_test_info()
        self.mismatch_index_dict = {}
            
        for mm in [0, 1, 2, 3, 4, 5, 6]:
            self.mismatch_index_dict[mm] = {}
            for f in self.fold_list:
                self.mismatch_index_dict[mm][f] = []
                sgrna_seq_list = self.fold_sgrna_info_dict["test_seq_list"][f]
                dataset_df_temp_list = []
                for sgrna_seq in sgrna_seq_list:
                    dataset_df_temp = self.dataset_df[self.dataset_df["sgRNA"] == sgrna_seq]
                    # Split offtarget or non-offtarget
                    data_df_sgrna_offtarget = dataset_df_temp[dataset_df_temp["reads"] >= 1]
                    data_df_sgrna_non_offtarget = dataset_df_temp[dataset_df_temp["reads"] == 0]
                    dataset_df_temp_list.append(data_df_sgrna_offtarget)
                    dataset_df_temp_list.append(data_df_sgrna_non_offtarget)
                dataset_df_temp = pd.concat(dataset_df_temp_list, axis=0)
                dataset_df_temp = dataset_df_temp.reset_index(drop=True)
                mismatch_index = dataset_df_temp[dataset_df_temp["Align.#Mismatches"] == mm].index.to_list()
                self.mismatch_index_dict[mm][f] += mismatch_index
        
        return self.mismatch_index_dict

    def return_sgrna_index_dict(self) -> dict:
        '''
        output:
        dict_type data
        sgrna_index_dict
        {fold: {sgrna_name: [start_index, end_index]}}
        '''
        if not self.flag_load_dataset:
            self.load_dataset()
        if not self.flag_fold_sgrna_info_dict:
            self.load_test_info()
        self.sgrna_index_dict = {}
        for f in self.fold_list:
            self.sgrna_index_dict[f] = {}
            sgrna_name_list = self.fold_sgrna_info_dict["test_names_list"][f]
            sgrna_seq_list = self.fold_sgrna_info_dict["test_seq_list"][f]
            start_index = 0
            for sgrna_name, sgrna_seq in zip(sgrna_name_list, sgrna_seq_list):
                dataset_df_sgrna = self.dataset_df[self.dataset_df["sgRNA"] == sgrna_seq]
                end_index = start_index + dataset_df_sgrna.shape[0]
                self.sgrna_index_dict[f][sgrna_name] = (start_index, end_index)
                end_index = start_index

        self.flag_sgrna_index_dict = True
        return self.sgrna_index_dict
    
    def aggregate_probabilities(self, ensemble: bool=True) -> dict:
        """
        Aggregates the probabilities from different models and folds, and optionally creates an ensemble of the best models.
        Parameters:
        ensemble (bool): If True, an ensemble of the best models based on PR-AUC metric will be created. Default is True.
        Returns:
        dict: A dictionary containing the aggregated probabilities.
              If datatype is changeseq or guideseq, the structure is:
              {model_name: {fold: {exp_id: probability numpy array}}}
              If ensemble is True, an additional key "ensemble" will be added with the same structure.
        """
        self.aggregated_probabilities_dict = {}
        for model_name in self.model_list:
            self.aggregated_probabilities_dict[model_name] = {}
            for f in self.fold_list:
                self.aggregated_probabilities_dict[model_name][f] = {}
                for e in self.exp_id_list:
                    probabilities_path = utilities_module.return_output_probability_path(model_name=model_name, datatype=self.datatype, fold=f, exp_id=e)
                    probabilities = np.load(probabilities_path)
                    self.aggregated_probabilities_dict[model_name][f][e] = probabilities
        
        # Find the best ensemble model and aggregate the probabilities
        if ensemble:
            best_ensemble_model_tuple = self.return_best_ensemble_models(metric="PR-AUC")
            self.aggregated_probabilities_dict["ensemble"] = {}
            for f in self.fold_list:
                self.aggregated_probabilities_dict["ensemble"][f] = {}
                for e in self.exp_id_list:
                    probabilities = np.zeros_like(self.aggregated_probabilities_dict[self.model_list[0]][f][e])
                    for model_name in best_ensemble_model_tuple:
                        probabilities += self.aggregated_probabilities_dict[model_name][f][e]
                    probabilities /= len(best_ensemble_model_tuple)
                    self.aggregated_probabilities_dict["ensemble"][f][e] = probabilities

        return self.aggregated_probabilities_dict
    
    
    def return_best_ensemble_models(self, metric: str) -> dict:
        # Calculate ensemble model score for each combination of model
        ensemble_score_dict = {}
        models_combinations_tuple_list = all_combinations(self.model_list)
        
        if self.datatype == "changeseq":
            for model_tuple in tqdm.tqdm(models_combinations_tuple_list, total=len(models_combinations_tuple_list), desc="Calculating ensemble model score"):
                ensemble_score_dict[model_tuple] = {"scores": []}
                for f in self.fold_list:
                    for e in self.exp_id_list:
                        probabilities = np.zeros_like(self.aggregated_probabilities_dict[model_tuple[0]][f][e])
                        for model_name in model_tuple:
                            probabilities += self.aggregated_probabilities_dict[model_name][f][e]
                        probabilities /= len(model_tuple)
                        true_label = self.test_label_dict[f]
                        for sgrna_name in self.fold_sgrna_info_dict["test_names_list"][f]:
                            sgrna_index = self.sgrna_index_dict[f][sgrna_name]
                            probabilities_sgrna = probabilities[sgrna_index[0]:sgrna_index[1]]
                            true_label_sgrna = true_label[sgrna_index[0]:sgrna_index[1]]
                            results = caluculate_metrics(true_label_sgrna, probabilities_sgrna, self.datatype, [metric])
                            ensemble_score_dict[model_tuple]["scores"].append(results[metric])
                ensemble_score_dict[model_tuple]["mean"] = np.mean(ensemble_score_dict[model_tuple]["scores"])
        
        if self.datatype == "guideseq" or self.datatype == "transfer":
            for model_tuple in tqdm.tqdm(models_combinations_tuple_list, total=len(models_combinations_tuple_list), desc="Calculating ensemble model score"):
                ensemble_score_dict[model_tuple] = {"scores": []}
                for f in self.fold_list:
                    for e in self.exp_id_list:
                        probabilities = np.zeros_like(self.aggregated_probabilities_dict[model_tuple[0]][f][e])
                        for model_name in model_tuple:
                            probabilities += self.aggregated_probabilities_dict[model_name][f][e]
                        probabilities /= len(model_tuple)
                        true_label = self.test_label_dict[f]
                        results = caluculate_metrics(true_label, probabilities, self.datatype, [metric])
                        ensemble_score_dict[model_tuple]["scores"].append(results[metric])
                ensemble_score_dict[model_tuple]["mean"] = np.mean(ensemble_score_dict[model_tuple]["scores"])
        
        # Return the best ensemble model
        self.best_ensemble_model_tuple = max(ensemble_score_dict, key=lambda x: ensemble_score_dict[x]["mean"])
        return self.best_ensemble_model_tuple
    
    
    def aggregate_result_by_metrics(self, ensemble: bool=True) -> dict:
        '''
        output:
        dict_type data
        result_dict
        if datatype is changeseq -> {model_name: {fold: {exp_id: {sgrna_name: {metrics: value}}}}}
        if datatype is guideseq -> {model_name: {fold: {exp_id: {metrics: value}}}}
        '''
        if ensemble:
            model_list_temp = self.model_list + ["ensemble"]
        else:
            model_list_temp = self.model_list
        
        self.result_dict = {}
        
        if self.datatype == "changeseq":
            if not self.flag_sgrna_index_dict:
                self.return_sgrna_index_dict()
            for model_name in model_list_temp:
                self.result_dict[model_name] = {}
                for f in self.fold_list:
                    self.result_dict[model_name][f] = {}
                    for e in self.exp_id_list:
                        self.result_dict[model_name][f][e] = {}
                        probabilities = self.aggregated_probabilities_dict[model_name][f][e]
                        true_label = self.test_label_dict[f]
                        for sgrna_name in self.fold_sgrna_info_dict["test_names_list"][f]:
                            sgrna_index = self.sgrna_index_dict[f][sgrna_name]
                            probabilities_sgrna = probabilities[sgrna_index[0]:sgrna_index[1]]
                            true_label_sgrna = true_label[sgrna_index[0]:sgrna_index[1]]
                            results = caluculate_metrics(true_label_sgrna, probabilities_sgrna, self.datatype, self.metrics_list)
                            self.result_dict[model_name][f][e][sgrna_name] = results
        
        if self.datatype == "guideseq" or self.datatype == "transfer":
            for model_name in model_list_temp:
                self.result_dict[model_name] = {}
                for f in self.fold_list:
                    self.result_dict[model_name][f] = {}
                    for e in self.exp_id_list:
                        probabilities = self.aggregated_probabilities_dict[model_name][f][e]
                        true_label = self.test_label_dict[f]
                        results = caluculate_metrics(true_label, probabilities, self.datatype, self.metrics_list)
                        self.result_dict[model_name][f][e] = results
        
        return self.result_dict
                        
    def whole_result(self, show: bool=False, to_excel: bool=False, ensemble: bool=True):
        """
        Generate and display or save the results of the model evaluation.
        Parameters:
        -----------
        show : bool, optional
            If True, prints the results to the console. Default is False.
        to_excel : bool, optional
            If True, saves the results to an Excel file. Default is False.
        ensemble : bool, optional
            If True, includes the ensemble model in the results. Default is True.
        Returns:
        --------
        None
        """
        
        if ensemble:
            model_list_temp = self.model_list + ["ensemble"]
        else:
            model_list_temp = self.model_list
        
        result_temp_dict = {}
        if self.datatype == "changeseq":
            for model_name in model_list_temp:
                result_temp_dict[model_name] = {}
                for f in self.fold_list:
                    for e in self.exp_id_list:
                        for sgrna_name in self.result_dict[model_name][f][e].keys():
                            for metrics in self.metrics_list:
                                if metrics in result_temp_dict[model_name].keys():
                                    if not np.isnan(self.result_dict[model_name][f][e][sgrna_name][metrics]):
                                        result_temp_dict[model_name][metrics].append(self.result_dict[model_name][f][e][sgrna_name][metrics])
                                else:
                                    if not np.isnan(self.result_dict[model_name][f][e][sgrna_name][metrics]):
                                        result_temp_dict[model_name][metrics] = [self.result_dict[model_name][f][e][sgrna_name][metrics]]
        
        if self.datatype == "guideseq" or self.datatype == "transfer":
            for model_name in model_list_temp:
                result_temp_dict[model_name] = {}
                for f in self.fold_list:
                    for e in self.exp_id_list:
                        for metrics in self.metrics_list:
                            if metrics in result_temp_dict[model_name].keys():
                                result_temp_dict[model_name][metrics].append(self.result_dict[model_name][f][e][metrics])
                            else:
                                result_temp_dict[model_name][metrics] = [self.result_dict[model_name][f][e][metrics]]
        # Print
        if show:
            for model_name in model_list_temp:
                print(f"Model: {self.model_name_mapping[model_name]}")
                result_statement_metrics = "\t"
                result_statement_mean = "mean:\t"
                result_statement_median = "median:\t"
                result_statement_std = "std:\t"
                result_statement_max = "max:\t"
                result_statement_min = "min:\t"
                for metrics in self.metrics_list:
                    result_array = np.array(result_temp_dict[model_name][metrics])
                    mean_result = np.mean(result_array)
                    median_result = np.median(result_array)
                    std_result = np.std(result_array)
                    max_result = np.max(result_array)
                    min_result = np.min(result_array)
                    if metrics in ["accuracy", "precision", "specificity"]:
                        result_statement_metrics += f"{metrics}\t"
                    else:
                        result_statement_metrics += f"{metrics}\t\t"
                    result_statement_mean += f"{mean_result:.4f}\t\t"
                    result_statement_median += f"{median_result:.4f}\t\t"
                    result_statement_std += f"{std_result:.4f}\t\t"
                    result_statement_max += f"{max_result:.4f}\t\t"
                    result_statement_min += f"{min_result:.4f}\t\t"
                print(result_statement_metrics)
                print(result_statement_mean)
                print(result_statement_median)
                print(result_statement_std)
                print(result_statement_max)
                print(result_statement_min)
                print("\n")
        
        if to_excel:
            wb = load_workbook(f"{config.result_base_dir_path}/{self.file_name}")
            ws_whole_result = wb.create_sheet("whole result")
            
            sheet_data = []
            if ensemble:
                sheet_data.append(["Ensemble model included: "] + [self.model_name_mapping[model_name] for model_name in self.best_ensemble_model_tuple])
            for model_name in model_list_temp:
                sheet_data.append([self.model_name_mapping[model_name]])
                metrics_list = ["metrics"]
                mean_list = ["mean"]
                median_list = ["median"]
                std_list = ["std"]
                max_list = ["max"]
                min_list = ["min"]
                for metrics in self.metrics_list:
                    metrics_list.append(self.metrics_mapping[metrics])
                    result_array = np.array(result_temp_dict[model_name][metrics])
                    mean_result = np.mean(result_array)
                    median_result = np.median(result_array)
                    std_result = np.std(result_array)
                    max_result = np.max(result_array)
                    min_result = np.min(result_array)
                    mean_list.append(mean_result)
                    median_list.append(median_result)
                    std_list.append(std_result)
                    max_list.append(max_result)
                    min_list.append(min_result)
                sheet_data.append(metrics_list)
                sheet_data.append(mean_list)
                sheet_data.append(median_list)
                sheet_data.append(std_list)
                sheet_data.append(max_list)
                sheet_data.append(min_list)
                sheet_data.append([])
            # Add sheet
            for row in sheet_data:
                ws_whole_result.append(row)
            wb.save(f"{config.result_base_dir_path}/{self.file_name}")
        return

    # 2 group non-parametric test
    def wilcoxon_signed_rank_test(self, show: bool=False, to_excel: bool=False, ensemble: bool=True):
        """
        Perform the Wilcoxon signed-rank test on model results and optionally display or save the results.
        Parameters:
        -----------
        show : bool, optional
            If True, print the Wilcoxon signed-rank test results to the console. Default is False.
        to_excel : bool, optional
            If True, save the Wilcoxon signed-rank test results to an Excel file. Default is False.
        ensemble : bool, optional
            If True, include the ensemble model in the comparison. Default is True.
        Returns:
        --------
        None
        Notes:
        ------
        - The function performs the Wilcoxon signed-rank test on the results of different models stored in `self.result_dict`.
        - The results are compared based on the metrics specified in `self.metrics_list`.
        - The function handles two types of data: "changeseq" and "guideseq".
        - For "changeseq" data, the results are aggregated over `sgrna_name`, `fold_list`, and `exp_id_list`.
        - For "guideseq" data, the results are aggregated over `fold_list` and `exp_id_list`.
        - The p-values from the Wilcoxon signed-rank test are stored in a dictionary and can be printed or saved to an Excel file.
        """
        
        if ensemble:
            model_list_temp = self.model_list + ["ensemble"]
        else:
            model_list_temp = self.model_list
        
        if self.datatype == "changeseq":
            # Store result of wilcoxon signed rank test
            wilcoxon_result_dict = {}
            for metrics in self.metrics_list:
                wilcoxon_result_dict[metrics] = {}
                for model_name_1 in model_list_temp:
                    for model_name_2 in model_list_temp:
                        pair_result_dict_temp = {model_name_1: [], model_name_2: []}
                        for f in self.fold_list:
                            for e in self.exp_id_list:
                                for sgrna_name in self.result_dict[model_name_1][f][e].keys():
                                    result1 = self.result_dict[model_name_1][f][e][sgrna_name][metrics]
                                    result2 = self.result_dict[model_name_2][f][e][sgrna_name][metrics]
                                    if not np.isnan(result1) and not np.isnan(result1):
                                        pair_result_dict_temp[model_name_1].append(result1)
                                    if not np.isnan(result2) and not np.isnan(result2):
                                        pair_result_dict_temp[model_name_2].append(result2)
                        # Wilcoxon signed rank test
                        w, p = wilcoxon(pair_result_dict_temp[model_name_1], pair_result_dict_temp[model_name_2], alternative='less')
                        wilcoxon_result_dict[metrics][(model_name_1, model_name_2)] = p
        
        if self.datatype == "guideseq" or self.datatype == "transfer":
            # Store result of wilcoxon signed rank test
            wilcoxon_result_dict = {}
            for metrics in self.metrics_list:
                wilcoxon_result_dict[metrics] = {}
                for model_name_1 in model_list_temp:
                    for model_name_2 in model_list_temp:
                        pair_result_dict_temp = {model_name_1: [], model_name_2: []}
                        for f in self.fold_list:
                            for e in self.exp_id_list:
                                result1 = self.result_dict[model_name_1][f][e][metrics]
                                result2 = self.result_dict[model_name_2][f][e][metrics]
                                pair_result_dict_temp[model_name_1].append(result1)
                                pair_result_dict_temp[model_name_2].append(result2)
                        # Wilcoxon signed rank test
                        w, p = wilcoxon(pair_result_dict_temp[model_name_1], pair_result_dict_temp[model_name_2], alternative='less')
                        wilcoxon_result_dict[metrics][(model_name_1, model_name_2)] = p
        
        # Print result
        if show:
            print("Wilcoxon signed rank test result")
            for metrics in self.metrics_list:
                print(f"Metrics: {metrics}")
                result_statement_fisrt_row = "\t\t\t"
                result_statement_after_first_row = []
                for model_name_1 in model_list_temp:
                    if model_name_1 in ["dnabert-no-pretrain", "crispr-bert-2025", "dnabert-epi-ablation"]:
                        result_statement_fisrt_row += f"{model_name_1}\t"
                        result_statenment_wilcoxon = f"{model_name_1}\t"
                    elif model_name_1 in ["dnabert-epi", "crispr-bert", "gru-embed", "crispr-hw", "crispr-dipoff", "ensemble"]:
                        result_statement_fisrt_row += f"{model_name_1}\t\t"
                        result_statenment_wilcoxon = f"{model_name_1}\t\t"
                    else:
                        result_statement_fisrt_row += f"{model_name_1}\t\t\t"
                        result_statenment_wilcoxon = f"{model_name_1}\t\t\t"
                    for model_name_2 in model_list_temp:
                        p_value = wilcoxon_result_dict[metrics][(model_name_1, model_name_2)]
                        if np.isnan(p_value):
                            result_statenment_wilcoxon += f"{p_value}\t\t\t"
                        else:
                            result_statenment_wilcoxon += f"{p_value:.3e}\t\t"
                    result_statement_after_first_row.append(result_statenment_wilcoxon)
                print(result_statement_fisrt_row)
                for result_statement in result_statement_after_first_row:
                    print(result_statement)
                print("\n")
            
        if to_excel:
            wb = load_workbook(f"{config.result_base_dir_path}/{self.file_name}")
            ws_wilcoxon_result = wb.create_sheet("wilcoxon result")
            
            sheet_data = []
            sheet_data.append(["Wilcoxon signed rank test result (p-value)"])
            sheet_data.append([])
            for metrics in self.metrics_list:
                sheet_data.append([self.metrics_mapping[metrics]])
                sheet_data.append([""] + [self.model_name_mapping[model_name] for model_name in model_list_temp])
                for model_name_1 in model_list_temp:
                    row_data = [self.model_name_mapping[model_name_1]]
                    for model_name_2 in model_list_temp:
                        p_value = wilcoxon_result_dict[metrics][(model_name_1, model_name_2)]
                        row_data.append(p_value)
                    sheet_data.append(row_data)
                sheet_data.append([])
            
            # Add sheet
            for row in sheet_data:
                ws_wilcoxon_result.append(row)
            wb.save(f"{config.result_base_dir_path}/{self.file_name}")
    
    def confusion_matrix(self, show: bool=False, to_excel: bool=False, ensemble: bool=True):
        """
        Generates and optionally displays or saves a confusion matrix for model predictions.
        Parameters:
        -----------
        show : bool, optional
            If True, prints the confusion matrix to the console. Default is False.
        to_excel : bool, optional
            If True, saves the confusion matrix to an Excel file. Default is False.
        ensemble : bool, optional
            If True, includes an ensemble model in the confusion matrix. Default is True.
        Returns:
        --------
        None
        Notes:
        ------
        - The function handles two types of data: "changeseq" and "guideseq".
        - For each type, it calculates the confusion matrix for different mismatch levels (0 to 6 and "all").
        - The confusion matrix includes counts of off-target and non-off-target predictions.
        - If `show` is True, the confusion matrix is printed to the console.
        - If `to_excel` is True, the confusion matrix is saved to an Excel file.
        """
        
        confusion_matrix_dict = {}
        
        if ensemble:
            model_list_temp = self.model_list + ["ensemble"]
        else:
            model_list_temp = self.model_list
        
        if self.datatype == "changeseq":
            for mm in ["all", 0, 1, 2, 3, 4, 5, 6]:
                confusion_matrix_dict[mm] = {}
                
                # Load true label
                confusion_matrix_dict[mm]["ground_truth"] = []
                for f in self.fold_list:
                    for e in self.exp_id_list:
                        if mm == "all":
                            confusion_matrix_dict[mm]["ground_truth"].append(
                                self.test_label_dict[f]
                            )
                        else:
                            confusion_matrix_dict[mm]["ground_truth"].append(
                                self.test_label_dict[f][self.mismatch_index_dict[mm][f]]
                            )
                confusion_matrix_dict[mm]["ground_truth"] = np.concatenate(confusion_matrix_dict[mm]["ground_truth"], axis=0)
                # Load probability and prediction
                for model_name in model_list_temp:
                    confusion_matrix_dict[mm][model_name] = {"probability": []}
                    for f in self.fold_list:
                        for e in self.exp_id_list:
                            if mm == "all":
                                confusion_matrix_dict[mm][model_name]["probability"].append(
                                    self.aggregated_probabilities_dict[model_name][f][e]
                                )
                            else:
                                confusion_matrix_dict[mm][model_name]["probability"].append(
                                    self.aggregated_probabilities_dict[model_name][f][e][self.mismatch_index_dict[mm][f]]
                                )
                    confusion_matrix_dict[mm][model_name]["probability"] = np.concatenate(confusion_matrix_dict[mm][model_name]["probability"], axis=0)
                    confusion_matrix_dict[mm][model_name]["prediction"] = np.argmax(confusion_matrix_dict[mm][model_name]["probability"], axis=1)
        
        if self.datatype == "guideseq" or self.datatype == "transfer":
            for mm in ["all", 0, 1, 2, 3, 4, 5, 6]:
                confusion_matrix_dict[mm] = {}
                confusion_matrix_dict[mm]["ground_truth"] = []
                for f in self.fold_list:
                    for e in self.exp_id_list:
                        if mm == "all":
                            confusion_matrix_dict[mm]["ground_truth"].append(self.test_label_dict[f])
                        else:
                            confusion_matrix_dict[mm]["ground_truth"].append(self.test_label_dict[f][self.mismatch_index_dict[mm][f]])
                confusion_matrix_dict[mm]["ground_truth"] = np.concatenate(confusion_matrix_dict[mm]["ground_truth"], axis=0)
                for model_name in model_list_temp:
                    confusion_matrix_dict[mm][model_name] = {"probability": []}
                    for f in self.fold_list:
                        for e in self.exp_id_list:
                            if mm == "all":
                                confusion_matrix_dict[mm][model_name]["probability"].append(self.aggregated_probabilities_dict[model_name][f][e])
                            else:
                                confusion_matrix_dict[mm][model_name]["probability"].append(self.aggregated_probabilities_dict[model_name][f][e][self.mismatch_index_dict[mm][f]])
                    confusion_matrix_dict[mm][model_name]["probability"] = np.concatenate(confusion_matrix_dict[mm][model_name]["probability"], axis=0)
                    confusion_matrix_dict[mm][model_name]["prediction"] = np.argmax(confusion_matrix_dict[mm][model_name]["probability"], axis=1)

        if show:
            for mm in ["all", 0, 1, 2, 3, 4, 5, 6]:
                print(f"mismatch: {mm}")
                # Count Off-target and Non Off-target
                ground_truth_array = confusion_matrix_dict[mm]["ground_truth"]
                offtarget_count = np.sum(ground_truth_array == 1)
                non_offtarget_count = np.sum(ground_truth_array == 0)
                print(f"off-target count: {offtarget_count}")
                print(f"non-off-target count: {non_offtarget_count}")
                
                # Confusion matrix
                for model_name in model_list_temp:
                    print(f"Model: {model_name}")
                    confusion_matrix = np.zeros((2, 2))
                    for i in range(2):
                        for j in range(2):
                            confusion_matrix[i, j] = np.sum((ground_truth_array == i) & (confusion_matrix_dict[mm][model_name]["prediction"] == j))
                    print("Ground Truth/Prediction\tNon-off-target\tOff-target")
                    print(f"Non-off-target\t\t{confusion_matrix[0, 0]:.4e}\t\t{confusion_matrix[0, 1]:.4e}")
                    print(f"Off-target\t\t{confusion_matrix[1, 0]:.4e}\t\t{confusion_matrix[1, 1]:.4e}")
                    print("\n")
        
        if to_excel:
            wb = load_workbook(f"{config.result_base_dir_path}/{self.file_name}")
            ws_confusion_matrix = wb.create_sheet("confusion matrix")
            
            for mm in ["all", 0, 1, 2, 3, 4, 5, 6]:
                sheet_data = []
                sheet_data.append([f"mismatch: {mm}"])
                # Count Off-target and Non Off-target
                ground_truth_array = confusion_matrix_dict[mm]["ground_truth"]
                offtarget_count = np.sum(ground_truth_array == 1)
                non_offtarget_count = np.sum(ground_truth_array == 0)
                sheet_data.append([f"off-target count: {offtarget_count}"])
                sheet_data.append([f"non-off-target count: {non_offtarget_count}"])
                sheet_data.append([])
                
                # Confusion matrix
                for model_name in model_list_temp:
                    sheet_data.append([f"Model: {self.model_name_mapping[model_name]}"])
                    confusion_matrix = np.zeros((2, 2))
                    for i in range(2):
                        for j in range(2):
                            confusion_matrix[i, j] = np.sum((ground_truth_array == i) & (confusion_matrix_dict[mm][model_name]["prediction"] == j))
                    sheet_data.append(["Ground Truth/Prediction", "Non-off-target", "Off-target"])
                    sheet_data.append(["Non-off-target", confusion_matrix[0, 0], confusion_matrix[0, 1]])
                    sheet_data.append(["Off-target", confusion_matrix[1, 0], confusion_matrix[1, 1]])
                    sheet_data.append([])
                # Add sheet
                for row in sheet_data:
                    ws_confusion_matrix.append(row)
            wb.save(f"{config.result_base_dir_path}/{self.file_name}")
                
                    
    def boxplot_for_result(self, ensemble: bool=False):
        
        if ensemble:
            model_list_temp = self.model_list + ["ensemble"]
        else:
            model_list_temp = self.model_list
        
        if not os.path.exists(f"{config.result_base_dir_path}/boxplot"):
            os.makedirs(f"{config.result_base_dir_path}/boxplot")
        
        result_temp_dict = {}
        if self.datatype == "changeseq":
            for model_name in model_list_temp:
                result_temp_dict[model_name] = {}
                for f in self.fold_list:
                    for e in self.exp_id_list:
                        for sgrna_name in self.result_dict[model_name][f][e].keys():
                            for metrics in self.metrics_list:
                                if metrics in result_temp_dict[model_name].keys():
                                    if not np.isnan(self.result_dict[model_name][f][e][sgrna_name][metrics]):
                                        result_temp_dict[model_name][metrics].append(self.result_dict[model_name][f][e][sgrna_name][metrics])
                                else:
                                    if not np.isnan(self.result_dict[model_name][f][e][sgrna_name][metrics]):
                                        result_temp_dict[model_name][metrics] = [self.result_dict[model_name][f][e][sgrna_name][metrics]]
        
        if self.datatype == "guideseq" or self.datatype == "transfer":
            for model_name in model_list_temp:
                result_temp_dict[model_name] = {}
                for f in self.fold_list:
                    for e in self.exp_id_list:
                        for metrics in self.metrics_list:
                            if metrics in result_temp_dict[model_name].keys():
                                result_temp_dict[model_name][metrics].append(self.result_dict[model_name][f][e][metrics])
                            else:
                                result_temp_dict[model_name][metrics] = [self.result_dict[model_name][f][e][metrics]]
        
        for metrics in self.metrics_list:
            boxplot_data = {}
            for model_name in model_list_temp:
                boxplot_data[model_name] = result_temp_dict[model_name][metrics]
            if "dnabert-epi" in model_list_temp:
                exist_dnabert_epi = True
            else:
                exist_dnabert_epi = False
            visualize_module.save_boxplot(
                data=boxplot_data,
                model_name_mapping=self.model_name_mapping,
                ylabel=self.metrics_mapping[metrics],
                save_path=f"{config.result_base_dir_path}/boxplot/boxplot_{self.datatype}_{metrics}",
                if_dnabert_epi=exist_dnabert_epi,
                dpi=300)
    
    
    def curve_roc_for_result(self, ensemble: bool=False):
        
        if ensemble:
            model_list_temp = self.model_list + ["ensemble"]
        else:
            model_list_temp = self.model_list
        # Create directory
        if not os.path.exists(f"{config.result_base_dir_path}/curve"):
            os.makedirs(f"{config.result_base_dir_path}/curve")
        
        # Load result
        # True label array
        ground_truth_label = []
        for f in self.fold_list:
            for e in self.exp_id_list:
                ground_truth_label.append(self.test_label_dict[f])
        ground_truth_label = np.concatenate(ground_truth_label, axis=0)
        # Probability array
        probability_dict = {}
        for model_name in model_list_temp:
            probability_dict[model_name] = []
            for f in self.fold_list:
                for e in self.exp_id_list:
                    probability_dict[model_name].append(self.aggregated_probabilities_dict[model_name][f][e])
            probability_dict[model_name] = np.concatenate(probability_dict[model_name], axis=0)
        # AUC-ROC value
        aucroc_dict = {}
        if self.datatype == "changeseq":
            for model_name in model_list_temp:
                roc_values = []
                prauc_values = []
                for f in self.fold_list:
                    for e in self.exp_id_list:
                        for sgrna_name in self.result_dict[model_name][f][e].keys():
                            if not np.isnan(self.result_dict[model_name][f][e][sgrna_name]["ROC-AUC"]):
                                roc_values.append(self.result_dict[model_name][f][e][sgrna_name]["ROC-AUC"])
                            prauc_values.append(self.result_dict[model_name][f][e][sgrna_name]["PR-AUC"])
                roc_value = np.mean(roc_values)
                prauc_value = np.mean(prauc_values)
                aucroc_dict[model_name] = {"ROC-AUC": np.round(roc_value, 4), "PR-AUC": np.round(prauc_value, 4)}
        if self.datatype == "guideseq" or self.datatype == "transfer":
            for model_name in model_list_temp:
                roc_values = []
                prauc_values = []
                for f in self.fold_list:
                    for e in self.exp_id_list:
                        roc_values.append(self.result_dict[model_name][f][e]["ROC-AUC"])
                        prauc_values.append(self.result_dict[model_name][f][e]["PR-AUC"])
                roc_value = np.mean(roc_values)
                prauc_value = np.mean(prauc_values)
                aucroc_dict[model_name] = {"ROC-AUC": np.round(roc_value, 4), "PR-AUC": np.round(prauc_value, 4)}
        
        # ROC curve
        # ground_truth_array: np.array, probability_dict: dict, rocauc_value_dict: dict, model_name_mapping: dict, title: str="", save_path: str=None, dpi: int=150
        visualize_module.plot_roc_curve(ground_truth_label, probability_dict, aucroc_dict, self.model_name_mapping, title="", save_path=f"{config.result_base_dir_path}/curve/roc_curve_{self.datatype}", dpi=300)
        visualize_module.plot_prroc_curve(ground_truth_label, probability_dict, aucroc_dict, self.model_name_mapping, title="", save_path=f"{config.result_base_dir_path}/curve/pr_curve_{self.datatype}", dpi=300)
    
    
    